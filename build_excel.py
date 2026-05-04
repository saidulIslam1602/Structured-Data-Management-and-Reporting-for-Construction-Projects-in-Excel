"""
Construction Project Data Management System v3
Advanced Excel formulas: COUNTIF/S, SUMIF/S, AVERAGEIF/S, VLOOKUP, XLOOKUP,
INDEX/MATCH, IF/IFS/IFERROR, conditional formatting, data validation, named ranges.

Design philosophy v3 — Mature / Executive:
  • Charcoal-black structure (one dark tone for ALL headers — no rainbow)
  • Single blue accent used sparingly
  • Muted pastel status badges (light bg + dark matching text — no solid bright fills)
  • Near-white alternating rows (barely perceptible)
  • Very thin light-gray borders throughout
  • Inspired by Bloomberg, McKinsey, and Deloitte Excel deliverables
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from datetime import date
import random

random.seed(42)

# ─── MATURE DESIGN TOKENS ────────────────────────────────────────────────────
# Philosophy: one dark structural color + one accent + neutral scale + muted badges

INK    = "0F172A"   # Slate-950 — sheet titles (near-black)
DARK   = "1E293B"   # Slate-800 — all column/section headers  (ONE color, no rainbow)
BODY   = "334155"   # Slate-700 — tertiary labels
MUTED  = "64748B"   # Slate-500 — notes, secondary text

WHITE  = "FFFFFF"   # Pure white — data cells, KPI values
SURF   = "F8FAFC"   # Slate-50  — alternating row tint  (barely visible, ~2% gray)
RULE   = "F1F5F9"   # Slate-100 — divider rows
EDGE   = "E2E8F0"   # Slate-200 — all cell borders  (very light)

ACC    = "2563EB"   # Blue-600  — THE single accent color
ACC_DK = "1D4ED8"   # Blue-700  — darker accent (for left-border on KPI cards)
ACC_LT = "EFF6FF"   # Blue-50   — accent tint for selected highlights

TEXT   = "0F172A"   # Slate-950 — primary body text

# ─── MUTED BADGE PALETTE ─────────────────────────────────────────────────────
# Each badge is (background, text) — all pastels with dark contrasting text
# This is the modern standard used by Linear, GitHub, Jira, Notion
B_GREEN  = ("DCFCE7", "14532D")   # Emerald-100 / Emerald-900
B_BLUE   = ("DBEAFE", "1E3A8A")   # Blue-100    / Blue-900
B_AMBER  = ("FEF3C7", "78350F")   # Amber-100   / Amber-900
B_RED    = ("FEE2E2", "7F1D1D")   # Red-100     / Red-900
B_GRAY   = ("F1F5F9", "334155")   # Slate-100   / Slate-700
B_TEAL   = ("CCFBF1", "134E4A")   # Teal-100    / Teal-900
B_PURPLE = ("F3E8FF", "581C87")   # Purple-100  / Purple-900

BADGE_MAP = {
    # Success-family
    "Completed":    B_GREEN,  "Closed":       B_GREEN,  "Resolved":     B_GREEN,
    "Ahead":        B_GREEN,  "Low":          B_GREEN,  "Environmental":B_GREEN,
    # Info-family
    "Open":         B_BLUE,   "Mitigated":    B_BLUE,   "On Track":     B_BLUE,
    "Material":     B_BLUE,   "Design":       B_BLUE,   "Technical":    B_BLUE,
    # Warning-family
    "In Progress":  B_AMBER,  "Medium":       B_AMBER,  "Workmanship":  B_AMBER,
    "Behind":       B_AMBER,  "Schedule":     B_AMBER,  "Quality":      B_AMBER,
    # Danger-family
    "Overdue":      B_RED,    "High":         B_RED,    "Critical":     B_RED,
    "Escalated":    B_RED,    "Safety":       B_RED,    "Behind_red":   B_RED,
    # Neutral-family
    "Not Started":  B_GRAY,   "On Hold":      B_GRAY,   "Cost":         B_GRAY,
    "Contractual":  B_GRAY,   "-":            B_GRAY,   "Stale Data":   B_GRAY,
    "Duplicate":    B_GRAY,   "Wrong Format": B_GRAY,
    # Teal-family
    "Installation": B_TEAL,
    # Validation error types
    "Missing Data": B_AMBER,  "Out of Range": B_RED,    "Logic Error":  B_RED,
}

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────
def F(c):
    return PatternFill("solid", fgColor=c)

def Fn(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def AL(h="center", v="center", wrap=True, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def thin_border():
    s = Side(style="thin", color=EDGE)
    return Border(left=s, right=s, top=s, bottom=s)

def accent_left_border(color=ACC):
    """Medium left border in accent color — used on KPI value cells."""
    return Border(
        left=Side(style="medium", color=color),
        right=Side(style="thin",  color=EDGE),
        top=Side(style="thin",    color=EDGE),
        bottom=Side(style="thin", color=EDGE),
    )

def header_cells(ws, row, cols, bg=DARK, size=10, height=26):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = F(bg)
        cell.font = Fn(bold=True, color=WHITE, size=size)
        cell.alignment = AL()
        cell.border = thin_border()
    ws.row_dimensions[row].height = height

def data_cells(ws, row, cols, alt=False):
    bg = SURF if alt else WHITE
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = F(bg)
        cell.font = Fn(bold=False, color=TEXT, size=10)
        cell.alignment = AL(h="left", indent=1)
        cell.border = thin_border()
    ws.row_dimensions[row].height = 20

def title_block(ws, text, cell_range, size=15, height=40):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = F(INK)
    c.font = Font(bold=True, color=WHITE, size=size, name="Calibri")
    c.alignment = AL()
    ws.row_dimensions[int("".join(filter(str.isdigit, cell_range.split(":")[0])))].height = height

def subtitle_block(ws, text, cell_range, height=22):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = F(DARK)
    c.font = Fn(bold=False, color="94A3B8", size=9)   # Slate-400 — subdued subtitle text
    c.alignment = AL(h="left", indent=2)
    ws.row_dimensions[int("".join(filter(str.isdigit, cell_range.split(":")[0])))].height = height

def section_hdr(ws, text, cell_range, size=10, height=24):
    """Uniform dark charcoal header — no per-section rainbow colors."""
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.fill = F(DARK)
    c.font = Fn(bold=True, color=WHITE, size=size)
    c.alignment = AL(h="left", indent=2)
    ws.row_dimensions[int("".join(filter(str.isdigit, cell_range.split(":")[0])))].height = height

def badge(ws, row, col, value):
    """Muted pastel status badge — light bg + dark text, never solid bright fill."""
    bg, tx = BADGE_MAP.get(value, B_GRAY)
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.fill = F(bg)
    cell.font = Font(bold=True, color=tx, size=9, name="Calibri")
    cell.alignment = AL()
    cell.border = thin_border()

def add_cf_badges(ws, range_str, anchor_col, start_row, statuses):
    """Conditional formatting using the muted badge palette."""
    for status in statuses:
        bg, tx = BADGE_MAP.get(status, B_GRAY)
        ws.conditional_formatting.add(
            range_str,
            FormulaRule(
                formula=[f'${anchor_col}{start_row}="{status}"'],
                fill=PatternFill("solid", fgColor=bg),
                font=Font(bold=True, color=tx, name="Calibri", size=9),
            ),
        )

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def print_setup(ws, landscape=True):
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.zoomScale   = 85

# ─── SAMPLE DATA ─────────────────────────────────────────────────────────────
DISCIPLINES = ["Civil", "Structural", "Electrical", "Mechanical",
               "Instrumentation", "Piping", "HVAC", "Scaffolding"]


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 0 — LOOKUP TABLES
# ══════════════════════════════════════════════════════════════════════════════
def build_lookup_tables(wb):
    ws = wb.create_sheet("Lookup Tables")
    ws.sheet_view.showGridLines = False

    title_block(ws, "LOOKUP TABLES — Reference Data for Formula Calculations", "A1:I1")
    subtitle_block(ws,
        "Powers VLOOKUP · XLOOKUP · INDEX/MATCH across all sheets. Do not rename or delete this sheet.",
        "A2:I2", height=22)

    # Table 1: Discipline → Lead Person
    section_hdr(ws, "TABLE 1 · Discipline → Lead Person  (VLOOKUP / XLOOKUP / INDEX-MATCH reference)", "A4:D4")
    header_cells(ws, 5, [1, 2, 3, 4], size=9)
    for col, hdr in zip([1, 2, 3, 4], ["Discipline", "Lead Person", "Dept. Code", "Email"]):
        ws.cell(row=5, column=col).value = hdr

    t1 = [
        ("Civil",           "J. Hansen",      "CIV", "j.hansen@project.com"),
        ("Structural",      "K. Lindqvist",   "STR", "k.lindqvist@project.com"),
        ("Electrical",      "M. Berg",        "ELE", "m.berg@project.com"),
        ("Mechanical",      "A. Nilsen",      "MEC", "a.nilsen@project.com"),
        ("Instrumentation", "T. Eriksson",    "INS", "t.eriksson@project.com"),
        ("Piping",          "S. Patel",       "PIP", "s.patel@project.com"),
        ("HVAC",            "L. Andersen",    "HVA", "l.andersen@project.com"),
        ("Scaffolding",     "R. Kowalski",    "SCA", "r.kowalski@project.com"),
        ("All",             "J. Hansen",      "ALL", "j.hansen@project.com"),
    ]
    for i, row in enumerate(t1):
        r = 6 + i
        data_cells(ws, r, [1, 2, 3, 4], alt=(i % 2 == 0))
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c).value = v

    # Table 2: Risk Score Bands
    section_hdr(ws, "TABLE 2 · Risk Score Reference  (IFS formula band reference)", "F4:I4")
    header_cells(ws, 5, [6, 7, 8, 9], size=9)
    for col, hdr in zip([6, 7, 8, 9], ["Min Score", "Max Score", "Level", "SLA Response"]):
        ws.cell(row=5, column=col).value = hdr

    t2 = [(15, 25, "High", "Immediate action"), (8, 14, "Medium", "Action within 14 days"), (1, 7, "Low", "Monitor monthly")]
    for i, (mn, mx, lvl, resp) in enumerate(t2):
        r = 6 + i
        data_cells(ws, r, [6, 7, 8, 9], alt=(i % 2 == 0))
        ws.cell(row=r, column=6).value = mn
        ws.cell(row=r, column=7).value = mx
        bg, tx = BADGE_MAP.get(lvl, B_GRAY)
        lc = ws.cell(row=r, column=8)
        lc.value = lvl; lc.fill = F(bg)
        lc.font = Font(bold=True, color=tx, size=9, name="Calibri"); lc.alignment = AL()
        ws.cell(row=r, column=9).value = resp

    # Table 3: Status Definitions
    section_hdr(ws, "TABLE 3 · Status Definitions & Dropdown Values", "A16:D16")
    header_cells(ws, 17, [1, 2, 3, 4], size=9)
    for col, hdr in zip([1, 2, 3, 4], ["Status", "Applies To", "Final?", "Description"]):
        ws.cell(row=17, column=col).value = hdr
    t3 = [
        ("Open",        "Actions / Risks / NCRs", "No",  "Raised, no action started"),
        ("In Progress", "Actions / NCRs",          "No",  "Work underway"),
        ("Closed",      "All",                     "Yes", "Fully resolved and verified"),
        ("Overdue",     "Actions",                 "No",  "Past due date — escalate"),
        ("Escalated",   "Risks",                   "No",  "Elevated to senior management"),
        ("Mitigated",   "Risks",                   "No",  "Mitigation complete; monitoring"),
        ("On Hold",     "Actions",                 "No",  "Blocked — awaiting input"),
    ]
    for i, (st, app, fin, desc) in enumerate(t3):
        r = 18 + i
        data_cells(ws, r, [1, 2, 3, 4], alt=(i % 2 == 0))
        badge(ws, r, 1, st)
        ws.cell(row=r, column=2).value = app
        ws.cell(row=r, column=3).value = fin
        ws.cell(row=r, column=4).value = desc

    # Table 4: Formula Index
    section_hdr(ws, "TABLE 4 · Formula Index — Where Each Advanced Function Is Used", "F16:I16")
    header_cells(ws, 17, [6, 7, 8, 9], size=9)
    for col, hdr in zip([6, 7, 8, 9], ["Sheet", "Column", "Function", "Purpose"]):
        ws.cell(row=17, column=col).value = hdr
    t4 = [
        ("Weekly Progress",   "Variance %",       "=H-G",       "Auto variance"),
        ("Weekly Progress",   "Status",            "IF nested",  "Auto status logic"),
        ("Weekly Progress",   "Responsible",       "XLOOKUP",    "From Lookup Tables"),
        ("Risk Register",     "Risk Score",        "=E×F",       "Auto P×I"),
        ("Risk Register",     "Risk Level",        "IF nested",  "From score"),
        ("Action Item Log",   "Responsible",       "VLOOKUP",    "From Lookup Tables"),
        ("Action Item Log",   "Days Overdue",      "IF/TODAY()", "Live days past due"),
        ("NCR Tracker",       "Days Open",         "IF/TODAY()", "Live days since raised"),
        ("NCR Tracker",       "Discipline Lead",   "INDEX/MATCH","From Lookup Tables"),
        ("Monthly Report",    "Schedule Var %",    "IFERROR/()", "Auto SV%"),
        ("Monthly Report",    "Cost Var %",        "IFERROR/()", "Auto CV%"),
        ("Monthly Report",    "Discipline Lead",   "INDEX/MATCH","From Lookup Tables"),
        ("DASHBOARD",         "All KPIs",          "COUNTIF",    "Live from data sheets"),
        ("DASHBOARD",         "Discipline table",  "AVERAGEIF",  "Live pull per discipline"),
        ("DASHBOARD",         "Monthly summary",   "SUMIFS",     "Multi-criteria totals"),
    ]
    for i, row in enumerate(t4):
        r = 18 + i
        data_cells(ws, r, [6, 7, 8, 9], alt=(i % 2 == 0))
        for c, v in enumerate(row, 6):
            ws.cell(row=r, column=c).value = v

    set_widths(ws, {"A": 18, "B": 18, "C": 12, "D": 28, "E": 2,
                    "F": 18, "G": 20, "H": 16, "I": 28})


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD  (all KPI values are live Excel formulas)
# ══════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.create_sheet("DASHBOARD")
    ws.sheet_view.showGridLines = False
    print_setup(ws, landscape=False)

    # Title
    title_block(ws, "CONSTRUCTION PROJECT MANAGEMENT DASHBOARD", "A1:O1", size=15, height=42)
    subtitle_block(ws,
        "[Site Name]  ·  [Project Name]  ·  [Client]  ·  Week: 18 / 2026  ·  All KPI values are live Excel formulas",
        "A2:O2", height=20)

    ws.row_dimensions[3].height = 12   # spacer

    # ── KPI TILES — Row 4 (label) + Row 5 (formula value) ────────────────────
    # Design: dark charcoal label row + white value cell + colored left border
    # Left border color signals sentiment: blue=neutral, charcoal=info, red=bad
    ws.row_dimensions[3].height = 10

    section_hdr(ws, "KEY PERFORMANCE INDICATORS", "A4:O4", size=10, height=22)

    kpis = [
        ("A", "OVERALL\nPROGRESS",
         "=IFERROR(AVERAGE('Weekly Progress'!$H$3:$H$40),0)",
         "0%", ACC, "AVERAGE(Actual%)"),

        ("D", "OPEN\nACTIONS",
         "=COUNTIF('Action Item Log'!$I$3:$I$100,\"Open\")"
         "+COUNTIF('Action Item Log'!$I$3:$I$100,\"Overdue\")",
         "0", "334155", "COUNTIF: Open + Overdue"),

        ("G", "OVERDUE\nACTIONS",
         "=COUNTIF('Action Item Log'!$I$3:$I$100,\"Overdue\")",
         "0", "7F1D1D", "COUNTIF: status=Overdue"),

        ("J", "OPEN / ESCALATED\nRISKS",
         "=COUNTIF('Risk Register'!$M$3:$M$100,\"Open\")"
         "+COUNTIF('Risk Register'!$M$3:$M$100,\"Escalated\")",
         "0", "78350F", "COUNTIF: Open+Escalated"),

        ("M", "OPEN\nNCRs",
         "=COUNTIF('NCR Quality Tracker'!$L$3:$L$100,\"Open\")"
         "+COUNTIF('NCR Quality Tracker'!$L$3:$L$100,\"In Progress\")",
         "0", "134E4A", "COUNTIF: Open+In Progress"),
    ]

    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 52
    ws.row_dimensions[7].height = 16

    for start_letter, label, formula, num_fmt, border_color, note in kpis:
        sc = ord(start_letter) - ord("A") + 1
        ec = sc + 2

        # Label row
        label_range = f"{start_letter}5:{get_column_letter(ec)}5"
        ws.merge_cells(label_range)
        lc = ws[f"{start_letter}5"]
        lc.value = label
        lc.fill = F(DARK)
        lc.font = Font(bold=True, color="94A3B8", size=8, name="Calibri")
        lc.alignment = AL()
        lc.border = thin_border()

        # Value row (formula cell)
        val_range = f"{start_letter}6:{get_column_letter(ec)}6"
        ws.merge_cells(val_range)
        vc = ws[f"{start_letter}6"]
        vc.value = formula
        vc.fill = F(WHITE)
        vc.font = Font(bold=True, color=INK, size=32, name="Calibri")
        vc.alignment = AL()
        vc.number_format = num_fmt
        vc.border = accent_left_border(border_color)

        # Note row
        note_range = f"{start_letter}7:{get_column_letter(ec)}7"
        ws.merge_cells(note_range)
        nc = ws[f"{start_letter}7"]
        nc.value = note
        nc.fill = F(SURF)
        nc.font = Font(italic=True, color=MUTED, size=7, name="Calibri")
        nc.alignment = AL()
        nc.border = thin_border()

    ws.row_dimensions[8].height = 12   # spacer

    # ── DISCIPLINE PROGRESS TABLE ─────────────────────────────────────────────
    section_hdr(ws, "DISCIPLINE PROGRESS  ·  Live via AVERAGEIF / COUNTIFS from Weekly Progress, Risk Register & Action Log",
                "A9:O9", size=10, height=22)

    disc_headers = ["Discipline", "Planned %", "Actual %", "Variance %", "Status",
                    "Open Actions", "Open Risks", "Open NCRs", "SPI  (Actual/Planned)"]
    header_cells(ws, 10, range(1, 10), size=9, height=28)
    for i, h in enumerate(disc_headers, 1):
        ws.cell(row=10, column=i).value = h

    for idx, disc in enumerate(DISCIPLINES):
        r = 11 + idx
        alt = idx % 2 == 0
        data_cells(ws, r, range(1, 10), alt=alt)
        ws.cell(row=r, column=1).value = disc

        # AVERAGEIF formulas
        ws.cell(row=r, column=2).value = f"=IFERROR(AVERAGEIF('Weekly Progress'!$B:$B,A{r},'Weekly Progress'!$G:$G),0)"
        ws.cell(row=r, column=2).number_format = "0%"
        ws.cell(row=r, column=3).value = f"=IFERROR(AVERAGEIF('Weekly Progress'!$B:$B,A{r},'Weekly Progress'!$H:$H),0)"
        ws.cell(row=r, column=3).number_format = "0%"
        ws.cell(row=r, column=4).value = f"=C{r}-B{r}"
        ws.cell(row=r, column=4).number_format = "+0%;-0%;0%"
        ws.cell(row=r, column=5).value = f'=IF(D{r}>0.03,"Ahead",IF(D{r}<-0.03,"Behind","On Track"))'

        # COUNTIFS
        ws.cell(row=r, column=6).value = (
            f'=COUNTIFS(\'Action Item Log\'!$E:$E,A{r},\'Action Item Log\'!$I:$I,"Open")'
            f'+COUNTIFS(\'Action Item Log\'!$E:$E,A{r},\'Action Item Log\'!$I:$I,"Overdue")'
        )
        ws.cell(row=r, column=7).value = (
            f'=COUNTIFS(\'Risk Register\'!$D:$D,A{r},\'Risk Register\'!$M:$M,"Open")'
            f'+COUNTIFS(\'Risk Register\'!$D:$D,A{r},\'Risk Register\'!$M:$M,"Escalated")'
        )
        ws.cell(row=r, column=8).value = (
            f'=COUNTIFS(\'NCR Quality Tracker\'!$B:$B,A{r},\'NCR Quality Tracker\'!$L:$L,"Open")'
            f'+COUNTIFS(\'NCR Quality Tracker\'!$B:$B,A{r},\'NCR Quality Tracker\'!$L:$L,"In Progress")'
        )
        ws.cell(row=r, column=9).value = f"=IFERROR(C{r}/B{r},0)"
        ws.cell(row=r, column=9).number_format = "0.00"

    # CF on Status column (col 5 = E)
    add_cf_badges(ws, "E11:E18", "E", 11, ["Ahead", "On Track", "Behind"])

    # Variance color scale
    ws.conditional_formatting.add("D11:D18", ColorScaleRule(
        start_type="min", start_color="FEE2E2",
        mid_type="num", mid_value=0, mid_color="F1F5F9",
        end_type="max", end_color="DCFCE7"
    ))

    # SPI color scale
    ws.conditional_formatting.add("I11:I18", ColorScaleRule(
        start_type="num", start_value=0.7, start_color="FEE2E2",
        mid_type="num",   mid_value=1.0,   mid_color="F1F5F9",
        end_type="num",   end_value=1.3,   end_color="DCFCE7"
    ))

    ws.row_dimensions[19].height = 12   # spacer

    # ── THREE SUMMARY PANELS ──────────────────────────────────────────────────
    # Risk (A-E), Action (G-K), NCR (M-O)

    section_hdr(ws, "RISK SUMMARY  ·  COUNTIFS", "A20:E20", height=22)
    header_cells(ws, 21, [1, 2, 3, 4, 5], size=9)
    for c, h in zip([1, 2, 3, 4, 5], ["Level", "Total", "Open", "Escalated", "Closed"]):
        ws.cell(row=21, column=c).value = h
    for i, lvl in enumerate(["High", "Medium", "Low"]):
        r = 22 + i
        data_cells(ws, r, [1, 2, 3, 4, 5], alt=(i % 2 == 0))
        badge(ws, r, 1, lvl)
        ws.cell(row=r, column=2).value = f'=COUNTIF(\'Risk Register\'!$H:$H,"{lvl}")'
        ws.cell(row=r, column=3).value = f'=COUNTIFS(\'Risk Register\'!$H:$H,"{lvl}",\'Risk Register\'!$M:$M,"Open")'
        ws.cell(row=r, column=4).value = f'=COUNTIFS(\'Risk Register\'!$H:$H,"{lvl}",\'Risk Register\'!$M:$M,"Escalated")'
        ws.cell(row=r, column=5).value = f'=COUNTIFS(\'Risk Register\'!$H:$H,"{lvl}",\'Risk Register\'!$M:$M,"Closed")'

    section_hdr(ws, "ACTION ITEM STATUS  ·  COUNTIF / COUNTIFS", "G20:K20", height=22)
    header_cells(ws, 21, [7, 8, 9, 10, 11], size=9)
    for c, h in zip([7, 8, 9, 10, 11], ["Status", "Count", "% of Total", "Critical/High", "Med/Low"]):
        ws.cell(row=21, column=c).value = h
    for i, st in enumerate(["Open", "In Progress", "Overdue", "Closed"]):
        r = 22 + i
        data_cells(ws, r, [7, 8, 9, 10, 11], alt=(i % 2 == 0))
        badge(ws, r, 7, st)
        ws.cell(row=r, column=8).value  = f'=COUNTIF(\'Action Item Log\'!$I:$I,"{st}")'
        ws.cell(row=r, column=9).value  = f'=IFERROR(H{r}/COUNTA(\'Action Item Log\'!$I$3:$I$100),0)'
        ws.cell(row=r, column=9).number_format = "0%"
        ws.cell(row=r, column=10).value = (
            f'=COUNTIFS(\'Action Item Log\'!$I:$I,"{st}",\'Action Item Log\'!$H:$H,"Critical")'
            f'+COUNTIFS(\'Action Item Log\'!$I:$I,"{st}",\'Action Item Log\'!$H:$H,"High")'
        )
        ws.cell(row=r, column=11).value = (
            f'=COUNTIFS(\'Action Item Log\'!$I:$I,"{st}",\'Action Item Log\'!$H:$H,"Medium")'
            f'+COUNTIFS(\'Action Item Log\'!$I:$I,"{st}",\'Action Item Log\'!$H:$H,"Low")'
        )

    section_hdr(ws, "NCR TYPE BREAKDOWN  ·  COUNTIFS", "M20:O20", height=22)
    header_cells(ws, 21, [13, 14, 15], size=9)
    for c, h in zip([13, 14, 15], ["NCR Type", "Total", "Open"]):
        ws.cell(row=21, column=c).value = h
    for i, typ in enumerate(["Material", "Workmanship", "Installation"]):
        r = 22 + i
        data_cells(ws, r, [13, 14, 15], alt=(i % 2 == 0))
        badge(ws, r, 13, typ)
        ws.cell(row=r, column=14).value = f'=COUNTIF(\'NCR Quality Tracker\'!$C:$C,"{typ}")'
        ws.cell(row=r, column=15).value = f'=COUNTIFS(\'NCR Quality Tracker\'!$C:$C,"{typ}",\'NCR Quality Tracker\'!$L:$L,"Open")'

    ws.row_dimensions[26].height = 12   # spacer

    # ── MONTHLY BUDGET SUMMARY (SUMIFS) ───────────────────────────────────────
    section_hdr(ws, "MONTHLY BUDGET & EARNED VALUE SUMMARY  ·  SUMIFS multi-criteria aggregation",
                "A27:O27", size=10, height=22)
    m_headers = ["Month", "Budget Hours\n(SUMIFS)", "Actual Hours\n(SUMIFS)", "Hours Var",
                 "Earned Value\n(SUMIFS)", "Planned Value\n(SUMIFS)", "SV  (EV−PV)", "CPI  (EV÷PV)"]
    header_cells(ws, 28, range(1, 9), size=9, height=32)
    for c, h in enumerate(m_headers, 1):
        ws.cell(row=28, column=c).value = h

    for i, month in enumerate(["Feb-26", "Mar-26", "Apr-26"]):
        r = 29 + i
        data_cells(ws, r, range(1, 9), alt=(i % 2 == 0))
        ws.cell(row=r, column=1).value = month
        ws.cell(row=r, column=2).value = f"=IFERROR(SUMIFS('Monthly Report'!$C:$C,'Monthly Report'!$A:$A,A{r}),0)"
        ws.cell(row=r, column=3).value = f"=IFERROR(SUMIFS('Monthly Report'!$D:$D,'Monthly Report'!$A:$A,A{r}),0)"
        ws.cell(row=r, column=4).value = f"=C{r}-B{r}"
        ws.cell(row=r, column=5).value = f"=IFERROR(SUMIFS('Monthly Report'!$G:$G,'Monthly Report'!$A:$A,A{r}),0)"
        ws.cell(row=r, column=5).number_format = '#,##0 "NOK"'
        ws.cell(row=r, column=6).value = f"=IFERROR(SUMIFS('Monthly Report'!$H:$H,'Monthly Report'!$A:$A,A{r}),0)"
        ws.cell(row=r, column=6).number_format = '#,##0 "NOK"'
        ws.cell(row=r, column=7).value = f"=E{r}-F{r}"
        ws.cell(row=r, column=7).number_format = '#,##0 "NOK"'
        ws.cell(row=r, column=8).value = f"=IFERROR(E{r}/F{r},0)"
        ws.cell(row=r, column=8).number_format = "0.000"

    # CF on Hours Var (col 4 = D)
    for col, rng in [("D", "D29:D31"), ("G", "G29:G31")]:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"${col}29<0"],
            fill=PatternFill("solid", fgColor="FEE2E2"),
            font=Font(bold=True, color="7F1D1D", name="Calibri", size=10)
        ))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"${col}29>=0"],
            fill=PatternFill("solid", fgColor="DCFCE7"),
            font=Font(bold=True, color="14532D", name="Calibri", size=10)
        ))

    ws.freeze_panes = "A10"
    set_widths(ws, {
        "A": 20, "B": 12, "C": 12, "D": 12, "E": 13, "F": 14,
        "G": 12, "H": 12, "I": 13, "J": 12, "K": 10,
        "L": 2,  "M": 14, "N": 10, "O": 10,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — WEEKLY PROGRESS  (XLOOKUP + formula-driven Status / Variance)
# ══════════════════════════════════════════════════════════════════════════════
def build_weekly_progress(wb):
    ws = wb.create_sheet("Weekly Progress")
    ws.sheet_view.showGridLines = False
    print_setup(ws)

    title_block(ws, "WEEKLY PROGRESS TRACKER — Construction Site", "A1:N1")
    subtitle_block(ws,
        "Manual entry: columns G (Planned %) and H (Actual %).  "
        "Formula-driven: I (Variance = H−G) · K (Status = IF logic) · L (Responsible = XLOOKUP) — do not overwrite.",
        "A2:N2", height=26)
    ws.row_dimensions[2].height = 26

    headers = [
        "ID", "Discipline", "Activity Description", "Location",
        "Planned\nStart", "Planned\nFinish",
        "Planned %\n↓ enter", "Actual %\n↓ enter",
        "Variance %\n=H−G", "Week",
        "Status\n=IF formula", "Responsible\n=XLOOKUP",
        "Submitted", "Comments"
    ]
    header_cells(ws, 3, range(1, 15), size=9, height=36)
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:N3"

    activities = [
        ("Civil",           "Excavation — Foundation Area B3",        "Zone B",    date(2026,3,10), date(2026,4,20), 1.00, 1.00, 18),
        ("Civil",           "Concrete Pour — Pile Caps Grid 4–7",     "Zone B",    date(2026,4,1),  date(2026,4,28), 0.80, 0.75, 18),
        ("Civil",           "Backfill & Compaction Block C",          "Zone C",    date(2026,4,15), date(2026,5,10), 0.40, 0.35, 18),
        ("Structural",      "Steel Erection — Main Frame Level 1",    "Building 1",date(2026,3,20), date(2026,5,5),  0.90, 0.95, 18),
        ("Structural",      "Secondary Steelwork — Mezzanine",        "Building 1",date(2026,4,10), date(2026,5,20), 0.55, 0.58, 18),
        ("Electrical",      "Cable Tray Installation — HV Zone",      "Substation",date(2026,4,1),  date(2026,5,15), 0.60, 0.55, 18),
        ("Electrical",      "LV Panel Wiring — Control Room",         "Ctrl Room", date(2026,4,20), date(2026,5,30), 0.25, 0.20, 18),
        ("Mechanical",      "Equipment Setting — Pump Skids P1–P4",   "Pump Hall", date(2026,4,5),  date(2026,5,10), 0.70, 0.72, 18),
        ("Mechanical",      "Crane Girder Installation",               "Warehouse", date(2026,4,15), date(2026,5,25), 0.40, 0.38, 18),
        ("Instrumentation", "Instrument Loop Drawing Review",          "All Areas", date(2026,3,25), date(2026,4,30), 0.95, 0.90, 18),
        ("Instrumentation", "Field Instrument Installation Ph.1",     "Zone A",    date(2026,4,20), date(2026,6,15), 0.20, 0.18, 18),
        ("Piping",          "Piping Spool Fabrication",               "Fab Shop",  date(2026,3,1),  date(2026,4,30), 1.00, 1.00, 18),
        ("Piping",          "Piping Installation — Process Area",     "Zone A",    date(2026,4,10), date(2026,6,1),  0.45, 0.48, 18),
        ("Piping",          "Hydrostatic Testing — Line P-101",       "Zone A",    date(2026,4,25), date(2026,5,5),  1.00, 1.00, 18),
        ("HVAC",            "Duct Fabrication & Delivery",            "Fab Shop",  date(2026,3,20), date(2026,4,25), 0.85, 0.80, 18),
        ("HVAC",            "Duct Installation — Ventilation Sys A",  "Building 1",date(2026,4,20), date(2026,5,30), 0.30, 0.25, 18),
        ("Scaffolding",     "Erect Scaffolding — Vessel V-201",       "Zone B",    date(2026,4,1),  date(2026,4,15), 1.00, 1.00, 18),
        ("Scaffolding",     "Scaffolding Maintenance Inspection",     "All",       date(2026,5,1),  date(2026,5,3),  1.00, 1.00, 18),
    ]

    last_data = 3 + len(activities)
    for idx, (disc, act, loc, ps, pf, plan, actual, week) in enumerate(activities):
        r = 4 + idx
        data_cells(ws, r, range(1, 15), alt=(idx % 2 == 0))
        ws.cell(row=r, column=1).value  = f"WP-{idx+1:03d}"
        ws.cell(row=r, column=2).value  = disc
        ws.cell(row=r, column=3).value  = act
        ws.cell(row=r, column=4).value  = loc
        for col, d in [(5, ps), (6, pf)]:
            dc = ws.cell(row=r, column=col); dc.value = d; dc.number_format = "DD-MMM-YY"
        ws.cell(row=r, column=7).value = plan;   ws.cell(row=r, column=7).number_format = "0%"
        ws.cell(row=r, column=8).value = actual; ws.cell(row=r, column=8).number_format = "0%"

        # FORMULA: Variance
        vc = ws.cell(row=r, column=9)
        vc.value = f"=H{r}-G{r}"; vc.number_format = "+0%;-0%;0%"

        ws.cell(row=r, column=10).value = week

        # FORMULA: Status — nested IF
        ws.cell(row=r, column=11).value = (
            f'=IF(H{r}>=1,"Completed",'
            f'IF(G{r}=0,"Not Started",'
            f'IF(H{r}-G{r}>0.03,"Ahead",'
            f'IF(G{r}-H{r}>0.03,"Behind","On Track"))))'
        )

        # FORMULA: Responsible — XLOOKUP
        ws.cell(row=r, column=12).value = (
            f"=IFERROR(XLOOKUP(B{r},"
            f"'Lookup Tables'!$A$6:$A$14,"
            f"'Lookup Tables'!$B$6:$B$14,"
            f'"TBD"),"-")'
        )

        sub = ws.cell(row=r, column=13); sub.value = date(2026, 5, 1); sub.number_format = "DD-MMM-YY"
        ws.cell(row=r, column=14).value = ""

    add_cf_badges(ws, f"K4:K{last_data}", "K", 4,
                  ["Completed", "Ahead", "On Track", "Behind", "Not Started"])

    # Actual % data bar (ACC blue)
    ws.conditional_formatting.add(f"H4:H{last_data}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=ACC))

    # Variance color scale (muted reds/greens)
    ws.conditional_formatting.add(f"I4:I{last_data}", ColorScaleRule(
        start_type="min", start_color="FEE2E2",
        mid_type="num", mid_value=0, mid_color="F1F5F9",
        end_type="max", end_color="DCFCE7"
    ))

    set_widths(ws, {
        "A": 10, "B": 17, "C": 42, "D": 16, "E": 12, "F": 12,
        "G": 11, "H": 11, "I": 11, "J": 7,  "K": 13, "L": 17,
        "M": 12, "N": 36,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — RISK REGISTER  (formula Score = P×I, Level = nested IF)
# ══════════════════════════════════════════════════════════════════════════════
def build_risk_register(wb):
    ws = wb.create_sheet("Risk Register")
    ws.sheet_view.showGridLines = False
    print_setup(ws)

    title_block(ws, "RISK REGISTER — Construction Project", "A1:N1")
    subtitle_block(ws,
        "Formula-driven: G (Score = Prob × Impact) · H (Level = IF nested) — enter only columns E and F to set risk level automatically.",
        "A2:N2", height=26)
    ws.row_dimensions[2].height = 26

    headers = [
        "Risk ID", "Category", "Risk Description", "Discipline",
        "Prob\n(1–5)", "Impact\n(1–5)",
        "Risk Score\n=E×F", "Risk Level\n=IF formula",
        "Mitigation Action", "Risk Owner",
        "Date Identified", "Review Date",
        "Status", "Comments"
    ]
    header_cells(ws, 3, range(1, 15), size=9, height=36)
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:N3"

    risks = [
        ("Safety",       "Working at height without valid scaffold inspection cert",         "Scaffolding",     4, 5, "Weekly scaffold inspection; tag all approved access points",               "K. Lindqvist",  date(2026,3,15), date(2026,5,15), "Open"),
        ("Schedule",     "Concrete delivery delays due to road restrictions",                "Civil",           3, 4, "Pre-order extra trucks; identify alternative delivery route",               "J. Hansen",     date(2026,3,20), date(2026,5,10), "Open"),
        ("Quality",      "Weld defects detected in structural steel connections",            "Structural",      2, 5, "Increase NDE scope to 20%; assign dedicated QC inspector",                 "M. Berg",       date(2026,4,1),  date(2026,5,20), "Escalated"),
        ("Technical",    "Instrument loop testing failure — signal noise on TI-101",        "Instrumentation", 3, 3, "Replace shielded cables; verify earthing configuration",                   "T. Eriksson",   date(2026,4,5),  date(2026,5,5),  "Open"),
        ("Cost",         "Material price escalation — copper cable +18%",                   "Electrical",      3, 3, "Lock prices via frame agreement; evaluate alternative suppliers",            "A. Nilsen",     date(2026,3,28), date(2026,5,28), "Open"),
        ("Environmental","Fuel spill risk near water body — Zone C",                        "Civil",           2, 4, "Install secondary containment; daily environmental officer check",          "K. Lindqvist",  date(2026,4,10), date(2026,5,30), "Open"),
        ("Schedule",     "Late delivery of critical pump P-201 from vendor",                "Mechanical",      4, 3, "Expedite vendor; activate contractual penalty clause if >1 week",          "A. Nilsen",     date(2026,4,12), date(2026,5,12), "Open"),
        ("Contractual",  "Design change — revised P&ID impacting piping route",             "Piping",          2, 3, "Issue formal change request; update schedule and cost estimate",            "S. Patel",      date(2026,4,18), date(2026,5,18), "Open"),
        ("Safety",       "Subcontractor personnel without site induction",                   "All",             2, 3, "Gate access control + daily induction records; revoke access on breach",   "L. Andersen",   date(2026,4,20), date(2026,6,1),  "Mitigated"),
        ("Technical",    "Fire suppression system pressure test failure",                   "Mechanical",      1, 4, "Pre-test with water; specialist commissioning team on standby",             "A. Nilsen",     date(2026,4,22), date(2026,6,15), "Open"),
        ("Quality",      "Painting rework due to humidity beyond 85% spec",                "Structural",      2, 2, "Monitor humidity; restrict painting to approved windows only",              "M. Berg",       date(2026,4,15), date(2026,5,15), "Mitigated"),
        ("Schedule",     "Permit delays for HV cable route crossing",                       "Electrical",      1, 2, "Engage authority early; submit permit 4 weeks in advance",                 "M. Berg",       date(2026,4,25), date(2026,6,1),  "Closed"),
        ("Cost",         "Scaffolding overhire duration beyond contract",                   "Scaffolding",     1, 2, "Weekly utilisation review; demobilise when work front closes",             "R. Kowalski",   date(2026,3,10), date(2026,5,1),  "Closed"),
        ("Technical",    "HVAC unit dimension mismatch with building opening",              "HVAC",            1, 3, "Review shop drawings vs. site measurement before delivery",                "L. Andersen",   date(2026,4,28), date(2026,5,28), "Open"),
        ("Environmental","Noise complaint from community during night work",                 "Civil",           1, 2, "Restrict noisy ops 07:00–22:00; notify community in advance",              "K. Lindqvist",  date(2026,4,30), date(2026,6,1),  "Open"),
    ]

    last_data = 3 + len(risks)
    for idx, risk in enumerate(risks):
        cat, desc, disc, prob, impact, mit, owner, id_d, rev_d, status = risk
        r = 4 + idx
        data_cells(ws, r, range(1, 15), alt=(idx % 2 == 0))
        ws.cell(row=r, column=1).value  = f"RSK-{idx+1:03d}"
        badge(ws, r, 2, cat)
        ws.cell(row=r, column=3).value  = desc
        ws.cell(row=r, column=4).value  = disc
        ws.cell(row=r, column=5).value  = prob
        ws.cell(row=r, column=6).value  = impact

        # FORMULA: Risk Score = Prob × Impact
        sc = ws.cell(row=r, column=7)
        sc.value = f"=E{r}*F{r}"; sc.alignment = AL()

        # FORMULA: Risk Level = nested IF
        lc = ws.cell(row=r, column=8)
        lc.value = f'=IF(G{r}=0,"-",IF(G{r}>=15,"High",IF(G{r}>=8,"Medium","Low")))'

        ws.cell(row=r, column=9).value  = mit
        ws.cell(row=r, column=10).value = owner
        for col, d in [(11, id_d), (12, rev_d)]:
            dc = ws.cell(row=r, column=col); dc.value = d; dc.number_format = "DD-MMM-YY"
        badge(ws, r, 13, status)
        ws.cell(row=r, column=14).value = ""

    # Risk Score heat map (muted reds/greens)
    ws.conditional_formatting.add(f"G4:G{last_data}", ColorScaleRule(
        start_type="num", start_value=1,  start_color="DCFCE7",
        mid_type="num",   mid_value=9,    mid_color="FEF3C7",
        end_type="num",   end_value=20,   end_color="FEE2E2"
    ))

    add_cf_badges(ws, f"H4:H{last_data}", "H", 4, ["High", "Medium", "Low"])
    add_cf_badges(ws, f"M4:M{last_data}", "M", 4, ["Open", "Mitigated", "Closed", "Escalated"])

    set_widths(ws, {
        "A": 10, "B": 16, "C": 50, "D": 17, "E": 8,  "F": 8,
        "G": 12, "H": 12, "I": 50, "J": 17, "K": 14, "L": 13,
        "M": 12, "N": 32,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ACTION ITEM LOG  (VLOOKUP + Days Overdue formula)
# ══════════════════════════════════════════════════════════════════════════════
def build_action_log(wb):
    ws = wb.create_sheet("Action Item Log")
    ws.sheet_view.showGridLines = False
    print_setup(ws)

    title_block(ws, "ACTION ITEM LOG — Site Coordination Meetings", "A1:N1")
    subtitle_block(ws,
        "F (Responsible) = VLOOKUP · N (Days Overdue) = IF(AND(≠Closed, TODAY()>DueDate), TODAY()−DueDate, \"\") · Status/Priority use dropdown validation",
        "A2:N2", height=26)
    ws.row_dimensions[2].height = 26

    headers = [
        "Action ID", "Source / Meeting", "Date Raised", "Action Description",
        "Discipline", "Responsible\n=VLOOKUP",
        "Due Date", "Priority", "Status",
        "% Done", "Date Closed", "Verified By", "Comments",
        "Days Overdue\n=IF/TODAY()"
    ]
    header_cells(ws, 3, range(1, 15), size=9, height=36)
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:N3"

    dv_status = DataValidation(type="list",
        formula1='"Open,In Progress,Closed,Overdue,On Hold"',
        showDropDown=False, showErrorMessage=True,
        errorTitle="Invalid", error="Choose from list")
    dv_priority = DataValidation(type="list",
        formula1='"Critical,High,Medium,Low"', showDropDown=False)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_priority)

    actions = [
        ("Site Coord W17",    date(2026,4,24), "Resolve concrete pour access conflict — Civil vs Piping Zone B",                    "Civil",           date(2026,4,28), "High",     "Closed",      1.0,  date(2026,4,27), "M. Berg"),
        ("Daily Standup W18", date(2026,4,27), "Submit updated cable installation schedule for approval by Project Engineer",       "Electrical",      date(2026,4,30), "High",     "Closed",      1.0,  date(2026,4,29), "T. Eriksson"),
        ("Quality Mtg W17",   date(2026,4,24), "Issue corrective action for weld rejection — NCR-007 structural connections",      "Structural",      date(2026,5,2),  "Critical", "In Progress", 0.60, None, ""),
        ("Site Coord W18",    date(2026,5,1),  "Update risk register with new schedule risks from vendor delays",                  "All",             date(2026,5,5),  "Medium",   "Open",        0.0,  None, ""),
        ("Client Mtg W17",    date(2026,4,23), "Prepare monthly progress report for Client review — May 2026",                    "All",             date(2026,5,7),  "High",     "Open",        0.0,  None, ""),
        ("Daily Standup W18", date(2026,4,28), "Calibrate pressure gauges before hydrostatic test — Piping Line P-102",            "Piping",          date(2026,4,30), "High",     "Closed",      1.0,  date(2026,4,30), "A. Nilsen"),
        ("Safety Mtg W18",    date(2026,4,29), "Conduct emergency evacuation drill before end of May",                            "All",             date(2026,5,20), "Medium",   "Open",        0.0,  None, "Coordinate HSE"),
        ("Quality Mtg W18",   date(2026,4,30), "Review and close outstanding punchlist items — Structural Level 1",               "Structural",      date(2026,5,9),  "Medium",   "In Progress", 0.30, None, ""),
        ("Site Coord W18",    date(2026,5,1),  "Confirm HVAC unit dimensions match building opening before delivery",              "HVAC",            date(2026,5,8),  "High",     "Open",        0.0,  None, ""),
        ("Client Mtg W17",    date(2026,4,23), "Issue revised programme baseline incorporating scope change SCR-012",             "All",             date(2026,5,3),  "Critical", "Overdue",     0.80, None, "Awaiting design"),
        ("Daily Standup W19", date(2026,5,4),  "Update SharePoint tracker — weekly discipline progress inputs for W19",           "All",             date(2026,5,6),  "Medium",   "Open",        0.0,  None, ""),
        ("Safety Mtg W17",    date(2026,4,22), "Replace damaged scaffolding clips — Tag Red Zone 4 out of service immediately",   "Scaffolding",     date(2026,4,24), "Critical", "Closed",      1.0,  date(2026,4,24), "L. Andersen"),
        ("Quality Mtg W17",   date(2026,4,24), "Submit material traceability certificates for steel batch SB-044",                "Structural",      date(2026,4,28), "High",     "Closed",      1.0,  date(2026,4,26), "M. Berg"),
        ("Site Coord W18",    date(2026,5,1),  "Validate instrument calibration data before loop testing",                        "Instrumentation", date(2026,5,12), "High",     "Open",        0.0,  None, ""),
        ("Client Mtg W18",    date(2026,4,30), "Confirm milestone dates — re-baseline if necessary; client sign-off required",    "All",             date(2026,5,14), "High",     "Open",        0.0,  None, ""),
        ("Daily Standup W18", date(2026,4,28), "Resolve interface clash — mechanical skid vs. electrical conduit Zone A",         "Mechanical",      date(2026,5,2),  "Medium",   "In Progress", 0.50, None, "3D review needed"),
        ("Safety Mtg W18",    date(2026,4,29), "Ensure all subcontractors complete updated Method Statement review",              "All",             date(2026,5,6),  "High",     "Open",        0.0,  None, ""),
    ]

    last_data = 3 + len(actions)
    for idx, action in enumerate(actions):
        source, d_raised, desc, disc, due, priority, status, pct, d_closed, verified = action
        r = 4 + idx
        data_cells(ws, r, range(1, 15), alt=(idx % 2 == 0))
        ws.cell(row=r, column=1).value = f"ACT-{idx+1:03d}"
        ws.cell(row=r, column=2).value = source
        dc = ws.cell(row=r, column=3); dc.value = d_raised; dc.number_format = "DD-MMM-YY"
        ws.cell(row=r, column=4).value = desc
        ws.cell(row=r, column=5).value = disc

        # FORMULA: Responsible = VLOOKUP
        ws.cell(row=r, column=6).value = (
            f"=IFERROR(VLOOKUP(E{r},'Lookup Tables'!$A$6:$B$14,2,0),\"-\")"
        )

        due_c = ws.cell(row=r, column=7); due_c.value = due; due_c.number_format = "DD-MMM-YY"
        badge(ws, r, 8, priority); dv_priority.add(ws.cell(row=r, column=8))
        badge(ws, r, 9, status);   dv_status.add(ws.cell(row=r, column=9))
        pct_c = ws.cell(row=r, column=10); pct_c.value = pct; pct_c.number_format = "0%"
        if d_closed:
            cc = ws.cell(row=r, column=11); cc.value = d_closed; cc.number_format = "DD-MMM-YY"
        ws.cell(row=r, column=12).value = verified
        ws.cell(row=r, column=13).value = ""

        # FORMULA: Days Overdue = IF(AND(not closed, past due), TODAY()-due, "")
        ws.cell(row=r, column=14).value = (
            f'=IF(AND(I{r}<>"Closed",I{r}<>"On Hold",'
            f'ISNUMBER(G{r}),TODAY()>G{r}),TODAY()-G{r},"")'
        )
        ws.cell(row=r, column=14).number_format = "0"

    add_cf_badges(ws, f"H4:H{last_data}", "H", 4, ["Critical", "High", "Medium", "Low"])
    add_cf_badges(ws, f"I4:I{last_data}", "I", 4, ["Open", "In Progress", "Closed", "Overdue", "On Hold"])

    # Days Overdue > 0 → muted red
    ws.conditional_formatting.add(f"N4:N{last_data}", FormulaRule(
        formula=["$N4>0"],
        fill=PatternFill("solid", fgColor="FEE2E2"),
        font=Font(bold=True, color="7F1D1D", name="Calibri", size=10)
    ))

    # % Done data bar
    ws.conditional_formatting.add(f"J4:J{last_data}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=ACC))

    set_widths(ws, {
        "A": 10, "B": 22, "C": 12, "D": 54, "E": 17, "F": 17,
        "G": 12, "H": 10, "I": 12, "J": 9,  "K": 12, "L": 15,
        "M": 30, "N": 13,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — NCR QUALITY TRACKER  (Days Open + INDEX/MATCH)
# ══════════════════════════════════════════════════════════════════════════════
def build_ncr_tracker(wb):
    ws = wb.create_sheet("NCR Quality Tracker")
    ws.sheet_view.showGridLines = False
    print_setup(ws)

    title_block(ws, "NON-CONFORMANCE REPORT (NCR) & QUALITY TRACKER", "A1:N1")
    subtitle_block(ws,
        "M (Days Open) = IF(Closed,0,TODAY()−DateRaised) · N (Discipline Lead) = INDEX/MATCH from Lookup Tables",
        "A2:N2", height=26)
    ws.row_dimensions[2].height = 26

    headers = [
        "NCR No.", "Discipline", "Type", "Non-Conformance Description",
        "Location", "Raised By", "Date Raised", "Root Cause",
        "Corrective Action", "Responsible", "Due Date", "Status",
        "Days Open\n=TODAY()−raised", "Disc. Lead\n=INDEX/MATCH"
    ]
    header_cells(ws, 3, range(1, 15), size=9, height=36)
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:N3"

    ncrs = [
        ("Structural",      "Material",    "Steel beam flange width 2mm below tolerance — spec 300mm, found 298mm",         "Zone B Grid 5-6",    "M. Berg",      date(2026,4,2),  "Wrong material ordered — spec not checked",                       "Replace beam; update procurement verification",                   "K. Lindqvist", date(2026,4,10), "Closed"),
        ("Civil",           "Workmanship", "Concrete cover to rebar 18mm vs. 40mm specified",                               "Foundation B3",      "J. Hansen",    date(2026,4,8),  "Formwork positioning error",                                      "Break out and reinstate; supervisor sign-off",                    "J. Hansen",    date(2026,4,18), "Closed"),
        ("Electrical",      "Installation","Cable tray support spacing 1200mm vs. 900mm per spec",                         "HV Room",            "A. Nilsen",    date(2026,4,12), "Incorrect reading of installation spec",                          "Add intermediate supports; re-inspect all trays",                 "M. Berg",      date(2026,4,22), "Closed"),
        ("Piping",          "Workmanship", "Porosity defect in butt weld — Line P-101 Spool 4",                            "Zone A Spool 4",     "S. Patel",     date(2026,4,15), "Welder qualification lapse; moisture in electrode",               "Cut out and re-weld; 100% NDE on 2 adjacent",                    "S. Patel",     date(2026,4,25), "Closed"),
        ("Structural",      "Workmanship", "Bolt torque not achieved — beam-to-column Level 1",                            "Building 1 L1",      "K. Lindqvist", date(2026,4,18), "Torque wrench not calibrated",                                    "Re-torque all; calibrate tools; 100% re-inspect",                "K. Lindqvist", date(2026,4,28), "Closed"),
        ("HVAC",            "Material",    "Duct insulation 25mm vs. 40mm specified",                                      "Building 1 L2",      "L. Andersen",  date(2026,4,20), "Incorrect material issued from stores",                           "Replace insulation; update material issue checklist",             "L. Andersen",  date(2026,5,1),  "In Progress"),
        ("Structural",      "Workmanship", "Incomplete fusion in fillet weld — Column C7",                                 "Zone C Col C7",      "M. Berg",      date(2026,4,22), "Welder fatigue and inadequate pre-heat",                          "Grind out and re-weld; PWHT review; briefing",                   "M. Berg",      date(2026,5,5),  "In Progress"),
        ("Civil",           "Workmanship", "Backfill compaction 91% Proctor vs. 95% minimum",                             "Zone C Layer 3",     "J. Hansen",    date(2026,4,25), "Compaction plant breakdown — work continued",                     "Remove and re-compact; equipment check first",                   "J. Hansen",    date(2026,5,7),  "Open"),
        ("Mechanical",      "Installation","Pump alignment out of tolerance — axial 0.12mm vs. 0.05mm max",               "Pump Hall P-203",    "A. Nilsen",    date(2026,4,28), "Grouting not cured before alignment",                             "Re-align after 28-day cure; document records",                   "A. Nilsen",    date(2026,5,10), "Open"),
        ("Instrumentation", "Installation","Wrong thermocouple type — type J installed vs. K specified",                  "Zone A TI-101",      "T. Eriksson",  date(2026,4,29), "Stores label incorrectly marked",                                 "Replace with type K; audit all IDs site-wide",                   "T. Eriksson",  date(2026,5,8),  "Open"),
        ("Piping",          "Workmanship", "Gasket material non-conformance — PTFE vs. spiral wound",                     "Zone A Flange F-22", "S. Patel",     date(2026,4,30), "Substitution without engineering approval",                       "Replace gaskets; issue material deviation form",                 "S. Patel",     date(2026,5,9),  "Open"),
        ("Electrical",      "Installation","Conduit penetration unsealed — missing firestop in firewall",                 "Control Room",       "M. Berg",      date(2026,5,1),  "Subcontractor unaware of firewall requirement",                   "Apply approved firestop; inspect all penetrations",              "M. Berg",      date(2026,5,12), "Open"),
    ]

    last_data = 3 + len(ncrs)
    for idx, ncr in enumerate(ncrs):
        disc, type_, desc, loc, raised_by, d_raised, root, corrective, responsible, due, status = ncr
        r = 4 + idx
        data_cells(ws, r, range(1, 15), alt=(idx % 2 == 0))
        ws.cell(row=r, column=1).value  = f"NCR-{idx+1:03d}"
        ws.cell(row=r, column=2).value  = disc
        badge(ws, r, 3, type_)
        ws.cell(row=r, column=4).value  = desc
        ws.cell(row=r, column=5).value  = loc
        ws.cell(row=r, column=6).value  = raised_by
        dr = ws.cell(row=r, column=7); dr.value = d_raised; dr.number_format = "DD-MMM-YY"
        ws.cell(row=r, column=8).value  = root
        ws.cell(row=r, column=9).value  = corrective
        ws.cell(row=r, column=10).value = responsible
        du = ws.cell(row=r, column=11); du.value = due; du.number_format = "DD-MMM-YY"
        badge(ws, r, 12, status)

        # FORMULA: Days Open
        ws.cell(row=r, column=13).value = f'=IF(L{r}="Closed",0,TODAY()-G{r})'
        ws.cell(row=r, column=13).number_format = "0"
        ws.cell(row=r, column=13).alignment = AL()

        # FORMULA: Discipline Lead = INDEX / MATCH
        ws.cell(row=r, column=14).value = (
            f"=IFERROR(INDEX('Lookup Tables'!$B$6:$B$14,"
            f"MATCH(B{r},'Lookup Tables'!$A$6:$A$14,0)),\"-\")"
        )

    add_cf_badges(ws, f"L4:L{last_data}", "L", 4, ["Open", "In Progress", "Closed"])
    add_cf_badges(ws, f"C4:C{last_data}", "C", 4, ["Material", "Workmanship", "Installation"])

    # Days Open > 28 → muted red, > 14 → muted amber
    ws.conditional_formatting.add(f"M4:M{last_data}", FormulaRule(
        formula=["$M4>28"], fill=PatternFill("solid", fgColor="FEE2E2"),
        font=Font(bold=True, color="7F1D1D", name="Calibri", size=10)))
    ws.conditional_formatting.add(f"M4:M{last_data}", FormulaRule(
        formula=["AND($M4>14,$M4<=28)"], fill=PatternFill("solid", fgColor="FEF3C7"),
        font=Font(bold=True, color="78350F", name="Calibri", size=10)))

    set_widths(ws, {
        "A": 10, "B": 17, "C": 14, "D": 52, "E": 20, "F": 13,
        "G": 12, "H": 44, "I": 44, "J": 16, "K": 12, "L": 13,
        "M": 12, "N": 17,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MONTHLY REPORT  (SUMIFS + SV/CV formulas + INDEX/MATCH)
# ══════════════════════════════════════════════════════════════════════════════
def build_monthly_report(wb):
    ws = wb.create_sheet("Monthly Report")
    ws.sheet_view.showGridLines = False
    print_setup(ws)

    title_block(ws, "MONTHLY PROGRESS REPORT — Input & Earned Value Analysis", "A1:O1")
    subtitle_block(ws,
        "Manual entry: A–H, K–M.  Formula-driven: I (SV%) · J (CV%) · N (Disc. Lead = INDEX/MATCH).  "
        "DASHBOARD SUMIFS pulls from cols A–H.",
        "A2:O2", height=26)
    ws.row_dimensions[2].height = 26

    headers = [
        "Month", "Discipline",
        "Budget\nHours", "Actual\nHours", "Planned\n%", "Actual\n%",
        "Earned Value\n(NOK)", "Planned Value\n(NOK)",
        "Schedule Var %\n=IFERROR((F−E)/E,0)",
        "Cost Var %\n=IFERROR((G−H)/H,0)",
        "Key Milestones", "Issues", "Actions\nRaised",
        "Disc. Lead\n=INDEX/MATCH", "Forecast\n%"
    ]
    header_cells(ws, 3, range(1, 16), size=9, height=44)
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:O3"

    monthly_data = [
        ("Feb-26","Civil",           1200,1180,0.30,0.28, 168000,180000,"Excavation complete Zone A",        "Soft ground — extra dewatering"),
        ("Feb-26","Structural",       800, 790,0.25,0.26,  97500, 93750,"Steel fabrication 100% complete",   ""),
        ("Feb-26","Electrical",       400, 380,0.15,0.14,  26600, 28000,"Substation foundation ready",       "Cable delivery delayed 1 week"),
        ("Mar-26","Civil",           1400,1450,0.55,0.53, 318000,330000,"Pile caps Zone A poured",           "Formwork shortage — mitigated"),
        ("Mar-26","Structural",      1200,1150,0.50,0.52, 312000,300000,"Level 1 steel erection complete",   ""),
        ("Mar-26","Electrical",       600, 580,0.30,0.28,  84000, 90000,"HV cable tray 60% installed",       ""),
        ("Mar-26","Mechanical",       500, 490,0.25,0.26,  65000, 62500,"Pump skids P1–P4 set",              ""),
        ("Mar-26","Piping",           800, 820,0.35,0.38, 152000,140000,"Fabrication complete; inst. started",""),
        ("Apr-26","Civil",           1600,1580,0.72,0.68, 408000,432000,"Zone B concrete pours started",     "2-day delay — road restrictions"),
        ("Apr-26","Structural",      1400,1390,0.65,0.70, 490000,455000,"Mezzanine steelwork in progress",   ""),
        ("Apr-26","Electrical",       900, 870,0.55,0.52, 234000,247500,"LV panel wiring started",           "NCR on cable tray spacing"),
        ("Apr-26","Mechanical",       700, 690,0.48,0.48, 168000,168000,"All pump equipment on site",        "Alignment NCR P-203"),
        ("Apr-26","Instrumentation",  500, 490,0.35,0.33,  82500, 87500,"Loop drawings reviewed — 80%",      "Thermocouple type error"),
        ("Apr-26","Piping",          1000, 980,0.60,0.62, 310000,300000,"Hydro test P-101 complete",         ""),
        ("Apr-26","HVAC",             600, 580,0.40,0.38, 114000,120000,"Ductwork 60% fabricated",           "Duct insulation NCR"),
        ("Apr-26","Scaffolding",      300, 295,0.80,0.82,  73800, 72000,"All active scaffolding inspected",  ""),
    ]

    last_data = 3 + len(monthly_data)
    for idx, row in enumerate(monthly_data):
        month, disc, bh, ah, pp, ap, ev, pv, milestones, issues = row
        r = 4 + idx
        data_cells(ws, r, range(1, 16), alt=(idx % 2 == 0))
        ws.cell(row=r, column=1).value  = month
        ws.cell(row=r, column=2).value  = disc
        ws.cell(row=r, column=3).value  = bh
        ws.cell(row=r, column=4).value  = ah
        ws.cell(row=r, column=5).value  = pp;  ws.cell(row=r, column=5).number_format  = "0.0%"
        ws.cell(row=r, column=6).value  = ap;  ws.cell(row=r, column=6).number_format  = "0.0%"
        ws.cell(row=r, column=7).value  = ev;  ws.cell(row=r, column=7).number_format  = '#,##0 "NOK"'
        ws.cell(row=r, column=8).value  = pv;  ws.cell(row=r, column=8).number_format  = '#,##0 "NOK"'

        # FORMULA: Schedule Variance %
        sv = ws.cell(row=r, column=9)
        sv.value = f"=IFERROR((F{r}-E{r})/E{r},0)"; sv.number_format = "+0.0%;-0.0%;0.0%"

        # FORMULA: Cost Variance %
        cv = ws.cell(row=r, column=10)
        cv.value = f"=IFERROR((G{r}-H{r})/H{r},0)"; cv.number_format = "+0.0%;-0.0%;0.0%"

        ws.cell(row=r, column=11).value = milestones
        ws.cell(row=r, column=12).value = issues
        ws.cell(row=r, column=13).value = random.randint(0, 3)

        # FORMULA: Discipline Lead = INDEX / MATCH
        ws.cell(row=r, column=14).value = (
            f"=IFERROR(INDEX('Lookup Tables'!$B$6:$B$14,"
            f"MATCH(B{r},'Lookup Tables'!$A$6:$A$14,0)),\"-\")"
        )
        ws.cell(row=r, column=15).value = min(1.0, ap + 0.06)
        ws.cell(row=r, column=15).number_format = "0%"

    # Muted CF on SV% and CV%
    for col_letter, rng in [("I", f"I4:I{last_data}"), ("J", f"J4:J{last_data}")]:
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"${col_letter}4<0"],
            fill=PatternFill("solid", fgColor="FEE2E2"),
            font=Font(bold=True, color="7F1D1D", name="Calibri", size=10)))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"${col_letter}4>0"],
            fill=PatternFill("solid", fgColor="DCFCE7"),
            font=Font(bold=True, color="14532D", name="Calibri", size=10)))

    # Actual % data bar
    ws.conditional_formatting.add(f"F4:F{last_data}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=ACC))

    set_widths(ws, {
        "A": 10, "B": 17, "C": 11, "D": 11, "E": 10, "F": 10,
        "G": 16, "H": 16, "I": 15, "J": 14, "K": 38,
        "L": 32, "M": 10, "N": 16, "O": 10,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — DATA VALIDATION LOG
# ══════════════════════════════════════════════════════════════════════════════
def build_data_validation_log(wb):
    ws = wb.create_sheet("Data Validation Log")
    ws.sheet_view.showGridLines = False
    print_setup(ws)

    title_block(ws, "DATA VALIDATION & QUALITY LOG", "A1:K1")
    ws.merge_cells("A2:K3")
    ic = ws["A2"]
    ic.value = (
        "PURPOSE: Log all data quality issues found during weekly/monthly validation sweeps. "
        "Every anomaly must be logged before correction to maintain a full audit trail. "
        "All Open items must be resolved before submission to client reporting."
    )
    ic.fill = F(SURF)
    ic.font = Font(italic=True, color=MUTED, size=9, name="Calibri")
    ic.alignment = AL(h="left", indent=2)
    ws.row_dimensions[2].height = 40

    headers = ["Log ID", "Date Found", "Data Source", "Field / Column",
               "Error Type", "Error Description", "Responsible",
               "Severity", "Action Taken", "Status", "Date Resolved"]
    header_cells(ws, 4, range(1, 12), size=9, height=26)
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i).value = h

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:K4"

    issues = [
        (date(2026,4,28),"Weekly Progress",  "Actual %",     "Missing Data",   "Actual % not entered for Instrumentation W17",                    "T. Eriksson","High",     "Followed up; data entered same day",                     "Resolved", date(2026,4,28)),
        (date(2026,4,28),"Risk Register",    "Review Date",  "Stale Data",     "6 risks with review date passed — not updated for 2+ weeks",      "S. Patel",  "Medium",   "All owners notified; register updated W18",              "Resolved", date(2026,4,29)),
        (date(2026,4,29),"Action Item Log",  "Due Date",     "Missing Data",   "3 actions with no due date assigned",                             "J. Hansen", "High",     "Dates confirmed in coord meeting W18",                   "Resolved", date(2026,4,29)),
        (date(2026,4,29),"Monthly Report",   "Actual Hours", "Out of Range",   "HVAC actual hours 5200 entered (budget 600) — data entry error",  "L. Andersen","Critical","Corrected to 580h; source timesheet confirmed",          "Resolved", date(2026,4,30)),
        (date(2026,4,30),"NCR Tracker",      "Root Cause",   "Missing Data",   "NCR-008, 009, 010 missing root cause entries",                    "M. Berg",   "Medium",   "Root causes added after investigation reports",          "Resolved", date(2026,5,1)),
        (date(2026,4,30),"Weekly Progress",  "Week No.",     "Wrong Format",   "Week entered as '18W' instead of integer 18",                     "R. Kowalski","Low",     "Corrected; format reminder sent to all leads",           "Resolved", date(2026,5,1)),
        (date(2026,5,1), "Monthly Report",   "EV Calc",      "Logic Error",    "Earned Value = −12000 for Piping Feb-26 — formula error",         "S. Patel",  "High",     "Formula corrected; cross-checked against cost plan",     "Resolved", date(2026,5,1)),
        (date(2026,5,2), "Action Item Log",  "Status",       "Stale Data",     "4 actions marked Open with due date passed",                      "J. Hansen", "High",     "Status updated to Overdue; owners notified",             "Resolved", date(2026,5,2)),
        (date(2026,5,3), "Risk Register",    "Risk Score",   "Logic Error",    "RSK-006 score = 0 — Probability and Impact were blank",           "S. Patel",  "Medium",   "Values entered; level updated to Medium (score = 8)",    "Resolved", date(2026,5,3)),
        (date(2026,5,4), "Weekly Progress",  "Responsible",  "Missing Data",   "2 activities have no Responsible person assigned",                "J. Hansen", "Medium",   "XLOOKUP auto-filled; manual review pending",             "Open",     None),
        (date(2026,5,4), "NCR Tracker",      "Date Closed",  "Logic Error",    "NCR-006 status=Closed but Date Closed cell is empty",             "M. Berg",   "High",     "Correction in progress",                                 "Open",     None),
        (date(2026,5,4), "Monthly Report",   "Forecast %",   "Out of Range",   "Forecast 112% for Civil Apr-26 — cannot exceed 100%",             "J. Hansen", "Medium",   "Capped at 100%; formula review in progress",             "Open",     None),
    ]

    severity_map = {"Critical": B_RED, "High": B_AMBER, "Medium": B_BLUE, "Low": B_GRAY}

    for idx, issue in enumerate(issues):
        d_found, source, field, err_type, desc, resp, severity, action, status, d_resolved = issue
        r = 5 + idx
        data_cells(ws, r, range(1, 12), alt=(idx % 2 == 0))
        ws.cell(row=r, column=1).value = f"VAL-{idx+1:03d}"
        df = ws.cell(row=r, column=2); df.value = d_found; df.number_format = "DD-MMM-YY"
        ws.cell(row=r, column=3).value  = source
        ws.cell(row=r, column=4).value  = field
        badge(ws, r, 5, err_type)
        ws.cell(row=r, column=6).value  = desc
        ws.cell(row=r, column=7).value  = resp
        bg, tx = severity_map.get(severity, B_GRAY)
        sc = ws.cell(row=r, column=8)
        sc.value = severity; sc.fill = F(bg)
        sc.font = Font(bold=True, color=tx, size=9, name="Calibri"); sc.alignment = AL()
        ws.cell(row=r, column=9).value  = action
        badge(ws, r, 10, status)
        if d_resolved:
            dr = ws.cell(row=r, column=11); dr.value = d_resolved; dr.number_format = "DD-MMM-YY"

    set_widths(ws, {
        "A": 10, "B": 12, "C": 20, "D": 16, "E": 16,
        "F": 50, "G": 14, "H": 10, "I": 46, "J": 11, "K": 13,
    })


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — INSTRUCTIONS & GUIDE
# ══════════════════════════════════════════════════════════════════════════════
def build_instructions(wb):
    ws = wb.create_sheet("Instructions & Guide")
    ws.sheet_view.showGridLines = False

    title_block(ws, "WORKBOOK GUIDE — Construction Project Data Management System", "A1:J1")
    subtitle_block(ws, "Sheet directory, formula reference, color legend, ID conventions, and usage workflow.",
                   "A2:J2", height=20)

    section_hdr(ws, "SHEET DIRECTORY", "A4:J4")
    header_cells(ws, 5, [1, 2, 3, 4, 5], size=9)
    for c, h in zip([1, 2, 3, 4, 5], ["Sheet", "Purpose", "Entry Columns", "Key Formulas", "Update Frequency"]):
        ws.cell(row=5, column=c).value = h
    sheets_info = [
        ("Lookup Tables",       "Reference tables — powers all lookup formulas",             "Update when personnel changes", "N/A",                "When team changes"),
        ("DASHBOARD",           "Live project health — KPIs, progress, risk, actions, EV",  "Read-only (formula-driven)",    "COUNTIF/S, AVERAGEIF, SUMIFS","Auto on file open"),
        ("Weekly Progress",     "Planned vs. actual % per activity per week",                "G (Planned%) and H (Actual%)",  "IF, XLOOKUP",        "Weekly — by Friday"),
        ("Risk Register",       "Risk log with P×I matrix and auto level",                   "All except G, H",               "E×F, nested IF",     "Weekly minimum"),
        ("Action Item Log",     "All meeting action items with status tracking",              "All except F, N",               "VLOOKUP, IF/TODAY()", "After each meeting"),
        ("NCR Quality Tracker", "Non-conformance full lifecycle",                             "All except M, N",               "TODAY()−date, INDEX/MATCH","As NCRs arise"),
        ("Monthly Report",      "Budget vs. actual + Earned Value analysis",                 "A–H, K–M",                      "SV/CV formulas, INDEX/MATCH, SUMIFS","2nd working day"),
        ("Data Validation Log", "Data quality issues audit trail",                           "All columns",                   "Manual log",         "During validation sweep"),
    ]
    for i, row in enumerate(sheets_info):
        r = 6 + i
        data_cells(ws, r, [1, 2, 3, 4, 5], alt=(i % 2 == 0))
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c).value = v

    section_hdr(ws, "FORMULA REFERENCE", "A16:J16")
    header_cells(ws, 17, [1, 2, 3, 4, 5], size=9)
    for c, h in zip([1, 2, 3, 4, 5], ["Function", "Sheet — Column", "Example", "Purpose", "Version"]):
        ws.cell(row=17, column=c).value = h
    formulas = [
        ("COUNTIF",    "Dashboard — KPI tiles",          '=COUNTIF($I:$I,"Open")',                          "Count matching single criterion",        "All"),
        ("COUNTIFS",   "Dashboard — discipline table",   '=COUNTIFS($E:$E,"Civil",$I:$I,"Open")',           "Count matching multiple criteria",       "2007+"),
        ("AVERAGEIF",  "Dashboard — discipline %",       '=AVERAGEIF($B:$B,"Civil",$H:$H)',                 "Average matching one criterion",         "2007+"),
        ("SUMIFS",     "Dashboard — monthly summary",    '=SUMIFS($C:$C,$A:$A,"Apr-26")',                   "Sum matching multiple criteria",         "2007+"),
        ("VLOOKUP",    "Action Log — Responsible col",   '=IFERROR(VLOOKUP(E3,$A$6:$B$14,2,0),"-")',        "Lookup value in first column of table",  "All"),
        ("XLOOKUP",    "Weekly Progress — Responsible",  '=IFERROR(XLOOKUP(B3,$A$6:$A$14,$B$6:$B$14),"?")',"Flexible; searches any direction",       "Excel 365"),
        ("INDEX/MATCH","NCR & Monthly — Disc. Lead",     '=INDEX($B$6:$B$14,MATCH(B3,$A$6:$A$14,0))',       "Flexible alternative to VLOOKUP",        "All"),
        ("IF nested",  "Weekly Progress — Status col",   '=IF(H3>=1,"Completed",IF(...))',                   "Multi-branch conditional logic",         "All"),
        ("IFERROR",    "All formula columns",             '=IFERROR(formula,0)',                              "Return fallback if formula errors",      "2007+"),
        ("TODAY()",    "Action Log — Days Overdue",       '=IF(AND(I3<>"Closed",TODAY()>G3),TODAY()-G3,"")', "Current date; updates on open",          "All"),
    ]
    for i, row in enumerate(formulas):
        r = 18 + i
        data_cells(ws, r, [1, 2, 3, 4, 5], alt=(i % 2 == 0))
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c).value = v

    section_hdr(ws, "COLOR & BADGE LEGEND", "A30:J30")
    legend = [
        (B_GREEN,  "Green badge",  "Completed · Closed · Resolved · Ahead · Low risk"),
        (B_BLUE,   "Blue badge",   "Open · On Track · Mitigated · Material"),
        (B_AMBER,  "Amber badge",  "In Progress · Medium · Workmanship · Warning"),
        (B_RED,    "Red badge",    "Overdue · High · Behind · Critical · Escalated"),
        (B_GRAY,   "Gray badge",   "Not Started · On Hold · Neutral / N/A"),
        (B_TEAL,   "Teal badge",   "Installation type NCRs"),
        ((SURF, TEXT), "Surface row", "Alternating data row background (barely visible)"),
        ((DARK, WHITE),"Dark header","Column headers and section headers — uniform charcoal"),
        ((INK, WHITE), "Ink title",  "Sheet titles — near-black"),
    ]
    for i, (colors, name, meaning) in enumerate(legend):
        r = 31 + i
        bg, tx = colors if isinstance(colors, tuple) else colors
        cc = ws.cell(row=r, column=1); cc.fill = F(bg); cc.border = thin_border()
        cc2 = ws.cell(row=r, column=2)
        cc2.value = name; cc2.fill = F(SURF)
        cc2.font = Font(bold=True, color=tx, size=9, name="Calibri")
        cc2.border = thin_border(); cc2.alignment = AL(h="left", indent=1)
        mc = ws.cell(row=r, column=3)
        mc.value = meaning; mc.fill = F(WHITE)
        mc.font = Fn(color=TEXT, size=9); mc.border = thin_border()
        mc.alignment = AL(h="left", indent=1)
        ws.row_dimensions[r].height = 20

    set_widths(ws, {"A": 18, "B": 30, "C": 55, "D": 44, "E": 14,
                    "F": 2, "G": 2, "H": 2, "I": 2, "J": 2})


# ══════════════════════════════════════════════════════════════════════════════
# ADVANCED EXCEL FEATURES
# Excel Tables · Data Validation · Row Grouping · Hyperlinks · Cell Comments
# ══════════════════════════════════════════════════════════════════════════════
def add_advanced_features(wb):
    """
    Adds:
      • Excel Tables  (native filter ↓ + sort on every column header)
      • Typed data validation  (dropdown, decimal range, integer range, date)
      • Column header comments  (hover tooltip explains each column)
      • Row grouping  (Monthly Report — collapse/expand by month)
      • Hyperlinks  (Dashboard discipline names → Weekly Progress)
      • Print titles  (header row repeats on every printed page)
    """

    # ── helpers ───────────────────────────────────────────────────────────────
    def excel_table(ws, ref, name):
        """Convert a range to an Excel Table — gives native sort/filter UI."""
        tab = Table(displayName=name, ref=ref)
        # TableStyleLight1 = minimal extra styling; our custom fills are preserved
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=False,  showColumnStripes=False,
        )
        ws.add_table(tab)

    def col_note(ws, row, col, title, body):
        """Hover comment on a header cell explaining the column purpose."""
        c = Comment(f"{title}\n{'─'*32}\n{body}", "Workbook Guide")
        c.width = 260; c.height = 90
        ws.cell(row=row, column=col).comment = c

    def dv_decimal(ws, sqref, prompt_title, prompt_body, lo=0.0, hi=1.0):
        dv = DataValidation(
            type="decimal", operator="between",
            formula1=str(lo), formula2=str(hi),
            allow_blank=True,
            showInputMessage=True, promptTitle=prompt_title, prompt=prompt_body,
            showErrorMessage=True, errorTitle="Out of Range",
            error=f"Enter a decimal between {lo} and {hi}  (e.g. 0.75 = 75%)"
        )
        ws.add_data_validation(dv); dv.sqref = sqref

    def dv_whole(ws, sqref, prompt_title, prompt_body, lo, hi):
        dv = DataValidation(
            type="whole", operator="between",
            formula1=str(lo), formula2=str(hi),
            allow_blank=True,
            showInputMessage=True, promptTitle=prompt_title, prompt=prompt_body,
            showErrorMessage=True, errorTitle="Invalid Number",
            error=f"Enter a whole number between {lo} and {hi}"
        )
        ws.add_data_validation(dv); dv.sqref = sqref

    def dv_list_range(ws, sqref, source, prompt_title, prompt_body):
        dv = DataValidation(
            type="list", formula1=source,
            allow_blank=True, showDropDown=False,
            showInputMessage=True, promptTitle=prompt_title, prompt=prompt_body,
            showErrorMessage=True, errorTitle="Invalid Selection",
            error="Choose a value from the dropdown list"
        )
        ws.add_data_validation(dv); dv.sqref = sqref

    def dv_date_after(ws, sqref, prompt_title, prompt_body):
        dv = DataValidation(
            type="date", operator="greaterThan",
            formula1="DATE(2020,1,1)",
            allow_blank=True,
            showInputMessage=True, promptTitle=prompt_title, prompt=prompt_body,
            showErrorMessage=True, errorTitle="Invalid Date",
            error="Enter a valid date after 01-Jan-2020 (format: DD-MMM-YY)"
        )
        ws.add_data_validation(dv); dv.sqref = sqref

    # ══════════════════════════════════════════════════════════════════════════
    # WEEKLY PROGRESS
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["Weekly Progress"]
    excel_table(ws, "A3:N21", "tblWeeklyProgress")

    # Discipline dropdown from Lookup Tables
    dv_list_range(ws, "B4:B21", "'Lookup Tables'!$A$6:$A$14",
                  "Discipline",
                  "Select from dropdown — list comes from Lookup Tables sheet")

    # Planned % and Actual % — must be 0.0 to 1.0
    dv_decimal(ws, "G4:G21", "Planned %",
               "Planned completion as decimal\n  0.25 = 25%\n  0.75 = 75%\n  1.00 = 100%")
    dv_decimal(ws, "H4:H21", "Actual %",
               "Actual completion as decimal\n  0.50 = 50%\n  1.00 = Complete")

    # Week number — 1 to 52
    dv_whole(ws, "J4:J21", "Week Number",
             "ISO week number\n  Range: 1 – 52\n  Current: Week 18", 1, 52)

    # Date columns
    dv_date_after(ws, "E4:E21", "Planned Start Date",
                  "Enter planned start date\n  Format: DD-MMM-YY\n  Example: 01-Apr-26")
    dv_date_after(ws, "F4:F21", "Planned Finish Date",
                  "Enter planned finish date\n  Must be after Planned Start")

    # Header comments on formula columns
    col_note(ws, 3, 9,  "Variance %  [FORMULA — do not edit]",
             "= Actual% − Planned%\n  Positive (+) = Ahead of schedule\n"
             "  Negative (−) = Behind schedule\n  Colour scale: red → white → green")
    col_note(ws, 3, 11, "Status  [FORMULA — do not edit]",
             "= IF(H≥1, 'Completed',\n    IF(H−G > 3%, 'Ahead',\n"
             "    IF(G−H > 3%, 'Behind', 'On Track')))\n"
             "  Muted badge colours applied via conditional formatting")
    col_note(ws, 3, 12, "Responsible  [XLOOKUP — auto-filled]",
             "= XLOOKUP(Discipline, LookupTables!A:A, LookupTables!B:B)\n"
             "  Auto-fills lead person from Lookup Tables.\n"
             "  You can override this cell manually if needed.")

    ws.print_title_rows = "3:3"    # repeat header row on every printed page

    # ══════════════════════════════════════════════════════════════════════════
    # RISK REGISTER
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["Risk Register"]
    excel_table(ws, "A3:N18", "tblRiskRegister")

    # Probability 1–5
    dv_whole(ws, "E4:E18", "Probability  (1 – 5)",
             "How likely is this risk to occur?\n"
             "  1 = Rare         (< 5% chance)\n"
             "  2 = Unlikely     (5–20%)\n"
             "  3 = Possible     (20–50%)\n"
             "  4 = Likely       (50–80%)\n"
             "  5 = Almost Certain (> 80%)", 1, 5)

    # Impact 1–5
    dv_whole(ws, "F4:F18", "Impact  (1 – 5)",
             "How severe is the impact if this risk occurs?\n"
             "  1 = Negligible  (minor inconvenience)\n"
             "  2 = Minor       (small delay / cost)\n"
             "  3 = Moderate    (1–2 week delay)\n"
             "  4 = Significant (>2 week delay)\n"
             "  5 = Critical    (project stoppage)", 1, 5)

    # Discipline dropdown
    dv_list_range(ws, "D4:D18", "'Lookup Tables'!$A$6:$A$14",
                  "Discipline", "Select affected discipline from dropdown")

    # Status dropdown
    dv_list_range(ws, "M4:M18", '"Open,Mitigated,Closed,Escalated"',
                  "Status", "Open / Mitigated / Closed / Escalated")

    # Date validation
    dv_date_after(ws, "K4:K18", "Date Identified",
                  "Date risk was formally identified\n  Format: DD-MMM-YY")
    dv_date_after(ws, "L4:L18", "Review Date",
                  "Next scheduled review date\n  Update at least weekly")

    # Header comments
    col_note(ws, 3, 7, "Risk Score  [FORMULA = E × F — do not edit]",
             "= Probability × Impact\n"
             "  1–7   = Low    (green heat-map)\n"
             "  8–14  = Medium (amber heat-map)\n"
             "  15–25 = High   (red heat-map)\n"
             "  Edit cols E and F; Score recalculates automatically.")
    col_note(ws, 3, 8, "Risk Level  [FORMULA — do not edit]",
             "= IF(Score≥15, 'High', IF(Score≥8, 'Medium', 'Low'))\n"
             "  Muted pastel badge applied automatically via conditional formatting.")

    ws.print_title_rows = "3:3"

    # ══════════════════════════════════════════════════════════════════════════
    # ACTION ITEM LOG
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["Action Item Log"]
    excel_table(ws, "A3:N20", "tblActionLog")

    dv_date_after(ws, "C4:C20", "Date Raised",
                  "Date action was formally raised\n  Format: DD-MMM-YY")
    dv_date_after(ws, "G4:G20", "Due Date",
                  "Target completion date\n  Used to calculate Days Overdue (col N)\n"
                  "  Format: DD-MMM-YY")
    dv_date_after(ws, "K4:K20", "Date Closed",
                  "Date action was resolved and verified\n  Format: DD-MMM-YY")

    dv_decimal(ws, "J4:J20", "% Done",
               "Completion % as decimal\n  0.0 = Not started\n  0.5 = 50% complete\n  1.0 = Done")

    dv_list_range(ws, "E4:E20", "'Lookup Tables'!$A$6:$A$14",
                  "Discipline", "Select responsible discipline from dropdown")

    dv_list_range(ws, "I4:I20", '"Open,In Progress,Closed,Overdue,On Hold"',
                  "Status", "Select current action status from dropdown")

    dv_list_range(ws, "H4:H20", '"Critical,High,Medium,Low"',
                  "Priority", "Select action priority from dropdown")

    col_note(ws, 3, 6,  "Responsible  [VLOOKUP — auto-filled]",
             "= VLOOKUP(Discipline, LookupTables, 2, 0)\n"
             "  Auto-fills the discipline lead from Lookup Tables.\n"
             "  Override manually if the action owner is a different person.")
    col_note(ws, 3, 14, "Days Overdue  [FORMULA — do not edit]",
             "= IF(AND(Status≠'Closed', Status≠'On Hold',\n"
             "         ISNUMBER(DueDate), TODAY() > DueDate),\n"
             "   TODAY() − DueDate, '')\n"
             "  Updates automatically each time the file is opened.\n"
             "  Red cell fill applied when value > 0.")

    ws.print_title_rows = "3:3"

    # ══════════════════════════════════════════════════════════════════════════
    # NCR QUALITY TRACKER
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["NCR Quality Tracker"]
    excel_table(ws, "A3:N15", "tblNCRTracker")

    dv_list_range(ws, "B4:B15", "'Lookup Tables'!$A$6:$A$14",
                  "Discipline", "Select discipline from dropdown")

    dv_list_range(ws, "C4:C15", '"Material,Workmanship,Installation,Design"',
                  "NCR Type",
                  "Material    = wrong material supplied\n"
                  "Workmanship = poor execution / craft\n"
                  "Installation = incorrectly fitted\n"
                  "Design       = design error")

    dv_list_range(ws, "L4:L15", '"Open,In Progress,Closed"',
                  "Status", "Open / In Progress / Closed")

    dv_date_after(ws, "G4:G15", "Date Raised",
                  "Date NCR was formally raised\n  Format: DD-MMM-YY")
    dv_date_after(ws, "K4:K15", "Due Date",
                  "Target resolution date\n  Format: DD-MMM-YY")

    col_note(ws, 3, 13, "Days Open  [FORMULA — do not edit]",
             "= IF(Status = 'Closed', 0, TODAY() − DateRaised)\n"
             "  Amber badge: > 14 days open\n"
             "  Red badge:   > 28 days open\n"
             "  Resets to 0 when Status is set to 'Closed'.")
    col_note(ws, 3, 14, "Discipline Lead  [INDEX/MATCH — auto-filled]",
             "= INDEX(LookupTables!B, MATCH(Discipline, LookupTables!A, 0))\n"
             "  Alternative to VLOOKUP — searches in any column.\n"
             "  Updates automatically if Discipline changes.")

    ws.print_title_rows = "3:3"

    # ══════════════════════════════════════════════════════════════════════════
    # MONTHLY REPORT — Row Grouping by Month (collapse / expand)
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["Monthly Report"]
    excel_table(ws, "A3:O19", "tblMonthlyReport")

    # Group rows by month block — click the [-] button on the left to collapse
    for r in range(4,  7):   ws.row_dimensions[r].outline_level = 1   # Feb-26
    for r in range(7,  12):  ws.row_dimensions[r].outline_level = 1   # Mar-26
    for r in range(12, 20):  ws.row_dimensions[r].outline_level = 1   # Apr-26
    ws.sheet_properties.outlinePr.summaryBelow = False   # group button above rows

    dv_list_range(ws, "B4:B19", "'Lookup Tables'!$A$6:$A$14",
                  "Discipline", "Select discipline from dropdown")
    dv_decimal(ws, "E4:E19", "Planned Progress %",
               "Planned completion % this month (decimal)\n  0.50 = 50%\n  1.00 = 100%")
    dv_decimal(ws, "F4:F19", "Actual Progress %",
               "Actual completion % achieved this month (decimal)")
    dv_whole(ws, "C4:C19", "Budgeted Hours",
             "Total budgeted man-hours for this discipline this month", 0, 99999)
    dv_whole(ws, "D4:D19", "Actual Hours",
             "Total actual man-hours spent this month", 0, 99999)

    col_note(ws, 3, 9,  "Schedule Variance %  [FORMULA — do not edit]",
             "= IFERROR((Actual% − Planned%) / Planned%, 0)\n"
             "  Green cell = ahead of schedule (positive)\n"
             "  Red cell   = behind schedule (negative)\n"
             "  Referenced by DASHBOARD SUMIFS monthly summary.")
    col_note(ws, 3, 10, "Cost Variance %  [FORMULA — do not edit]",
             "= IFERROR((EV − PV) / PV, 0)\n"
             "  Green = EV > PV (under budget)\n"
             "  Red   = EV < PV (over budget)\n"
             "  CPI chart on DASHBOARD uses Dashboard SUMIFS of this data.")
    col_note(ws, 3, 14, "Discipline Lead  [INDEX/MATCH — auto-filled]",
             "= INDEX(LookupTables!B:B, MATCH(Discipline, LookupTables!A:A, 0))\n"
             "  Updates automatically when Discipline cell changes.")

    ws.print_title_rows = "3:3"

    # ══════════════════════════════════════════════════════════════════════════
    # DATA VALIDATION LOG
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["Data Validation Log"]
    excel_table(ws, "A4:K16", "tblValidationLog")
    dv_date_after(ws, "B5:B16", "Date Found",
                  "Date the data quality issue was identified\n  Format: DD-MMM-YY")
    dv_date_after(ws, "K5:K16", "Date Resolved",
                  "Date the issue was corrected and verified\n  Format: DD-MMM-YY")
    ws.print_title_rows = "4:4"

    # ══════════════════════════════════════════════════════════════════════════
    # DASHBOARD — Hyperlinks from discipline names to Weekly Progress sheet
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb["DASHBOARD"]
    for r in range(11, 19):
        cell = ws.cell(row=r, column=1)
        if cell.value:
            cell.hyperlink = "#'Weekly Progress'!A3"
            cell.font = Font(color=ACC, bold=False, size=10,
                             name="Calibri", underline="single")

    print("Advanced features added:")
    print("  • Excel Tables with native sort/filter on all 6 data sheets")
    print("  • Typed data validation: dropdown, decimal, integer, date")
    print("  • Column header hover comments on all formula columns")
    print("  • Row grouping by month on Monthly Report (collapse / expand)")
    print("  • Hyperlinks: Dashboard discipline names → Weekly Progress")
    print("  • Print titles: header row repeats on every printed page")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    wb.remove(wb.active)

    print("Building Lookup Tables..."); build_lookup_tables(wb)
    print("Building Dashboard...");    build_dashboard(wb)
    print("Building Weekly Progress..."); build_weekly_progress(wb)
    print("Building Risk Register..."); build_risk_register(wb)
    print("Building Action Item Log..."); build_action_log(wb)
    print("Building NCR Quality Tracker..."); build_ncr_tracker(wb)
    print("Building Monthly Report..."); build_monthly_report(wb)
    print("Building Data Validation Log..."); build_data_validation_log(wb)
    print("Building Instructions & Guide..."); build_instructions(wb)

    # Charts
    print("Adding charts...")
    add_charts(wb)

    # Advanced features: Tables, Validation, Comments, Grouping, Hyperlinks
    print("Adding advanced features...")
    add_advanced_features(wb)

    # Named ranges
    wb.defined_names.add(DefinedName("DisciplineLeadTable", attr_text="'Lookup Tables'!$A$6:$B$14"))
    wb.defined_names.add(DefinedName("DisciplineList",      attr_text="'Lookup Tables'!$A$6:$A$14"))

    # Tab colors — two tones only: ink for data sheets, accent for reference
    tab_map = {
        "DASHBOARD":           INK,
        "Weekly Progress":     DARK,
        "Risk Register":       DARK,
        "Action Item Log":     DARK,
        "NCR Quality Tracker": DARK,
        "Monthly Report":      DARK,
        "Data Validation Log": DARK,
        "Lookup Tables":       BODY,
        "Instructions & Guide":BODY,
    }
    for name, color in tab_map.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    # Sheet order
    order = ["DASHBOARD", "Weekly Progress", "Risk Register", "Action Item Log",
             "NCR Quality Tracker", "Monthly Report", "Data Validation Log",
             "Lookup Tables", "Instructions & Guide"]
    sheet_objects = {s.title: s for s in wb._sheets}
    wb._sheets = [sheet_objects[n] for n in order if n in sheet_objects]
    wb.active = wb["DASHBOARD"]

    output = (
        "/home/saidul/Desktop/"
        "Structured Data Management and Reporting for Construction Projects in Excel/"
        "Construction_Project_Data_Management.xlsx"
    )
    wb.save(output)
    print(f"\nSaved: {output}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

# ══════════════════════════════════════════════════════════════════════════════
# CHARTS — added after all sheets are built so cross-sheet references work
# ══════════════════════════════════════════════════════════════════════════════
def add_charts(wb):
    """
    4 charts across 3 sheets:
      1. DASHBOARD     — Discipline Progress: Planned vs Actual % (clustered column)
      2. DASHBOARD     — Schedule Performance Index by Discipline (horizontal bar)
      3. Weekly Progress — Activity Progress: Planned vs Actual (horizontal bar)
      4. Monthly Report  — EV vs PV by Month (clustered column) + SV% line
    """

    def bar_line(series, hex_color, width_emu=25000):
        """Set solid fill on a bar series and outline."""
        series.graphicalProperties.solidFill = hex_color
        series.graphicalProperties.line.solidFill = hex_color

    def line_style(series, hex_color, width_emu=25000):
        """Set line color and width for a LineChart series."""
        series.graphicalProperties.line.solidFill = hex_color
        series.graphicalProperties.line.width = width_emu

    def make_bar(title, chart_type="col", grouping="clustered",
                 w=18, h=12, y_fmt=None, x_fmt=None, legend_pos="b"):
        c = BarChart()
        c.type      = chart_type
        c.grouping  = grouping
        c.style     = 10          # Excel minimal clean style
        c.title     = title
        c.width     = w
        c.height    = h
        if y_fmt:
            c.y_axis.numFmt = y_fmt
        if x_fmt:
            c.x_axis.numFmt = x_fmt
        c.legend.position = legend_pos
        # Remove chart border / plot-area border for clean look
        c.plot_area.graphicalProperties = None
        return c

    # ── Palette shortcuts ─────────────────────────────────────────────────────
    SLATE  = "94A3B8"   # muted slate  — "planned" / secondary series
    BLUE   = "2563EB"   # accent blue  — "actual" / primary series
    CHARCOAL = "1E293B" # dark charcoal — single-series charts

    # ══════════════════════════════════════════════════════════════════════════
    # CHART 1 — DASHBOARD: Discipline Progress (Planned vs Actual %)
    # ══════════════════════════════════════════════════════════════════════════
    ws_dash = wb["DASHBOARD"]

    section_hdr(ws_dash,
        "VISUAL ANALYTICS  ·  Charts refresh automatically when data changes",
        "A34:O34", height=22)

    c1 = make_bar("Discipline Progress — Planned vs Actual %",
                  chart_type="col", y_fmt="0%", w=18, h=12)
    c1.y_axis.title = "Progress %"
    c1.y_axis.scaling.min = 0

    cats1   = Reference(ws_dash, min_col=1, min_row=11, max_row=18)
    planned = Reference(ws_dash, min_col=2, min_row=10, max_row=18)
    actual  = Reference(ws_dash, min_col=3, min_row=10, max_row=18)
    c1.add_data(planned, titles_from_data=True)
    c1.add_data(actual,  titles_from_data=True)
    bar_line(c1.series[0], SLATE)   # Planned — muted
    bar_line(c1.series[1], BLUE)    # Actual  — accent
    c1.set_categories(cats1)
    ws_dash.add_chart(c1, "A35")

    # ══════════════════════════════════════════════════════════════════════════
    # CHART 2 — DASHBOARD: Schedule Performance Index (SPI) by Discipline
    # ══════════════════════════════════════════════════════════════════════════
    c2 = make_bar("Schedule Performance Index (SPI)  ·  Target = 1.00",
                  chart_type="bar", x_fmt="0.00", w=14, h=12, legend_pos="b")
    c2.x_axis.title = "SPI  (Actual ÷ Planned)"

    cats2    = Reference(ws_dash, min_col=1, min_row=11, max_row=18)
    spi_data = Reference(ws_dash, min_col=9, min_row=10, max_row=18)
    c2.add_data(spi_data, titles_from_data=True)
    bar_line(c2.series[0], CHARCOAL)
    c2.set_categories(cats2)
    ws_dash.add_chart(c2, "J35")

    # ══════════════════════════════════════════════════════════════════════════
    # CHART 3 — WEEKLY PROGRESS: Activity Progress (Planned vs Actual %)
    # ══════════════════════════════════════════════════════════════════════════
    ws_wp = wb["Weekly Progress"]

    section_hdr(ws_wp,
        "ACTIVITY PROGRESS CHART  ·  Planned vs Actual % per Activity",
        "A23:N23", height=22)

    c3 = make_bar("Activity Progress — Planned vs Actual %",
                  chart_type="bar", x_fmt="0%", w=24, h=18)
    c3.x_axis.title = "Progress %"

    cats3   = Reference(ws_wp, min_col=3, min_row=4,  max_row=21)  # Activity names
    plan_wp = Reference(ws_wp, min_col=7, min_row=3,  max_row=21)  # Planned % + header
    act_wp  = Reference(ws_wp, min_col=8, min_row=3,  max_row=21)  # Actual %  + header
    c3.add_data(plan_wp, titles_from_data=True)
    c3.add_data(act_wp,  titles_from_data=True)
    bar_line(c3.series[0], SLATE)
    bar_line(c3.series[1], BLUE)
    c3.set_categories(cats3)
    ws_wp.add_chart(c3, "A24")

    # ══════════════════════════════════════════════════════════════════════════
    # CHART 4 — MONTHLY REPORT: EV vs PV by Month (clustered column)
    # Pulls from DASHBOARD rows 29–31 (SUMIFS monthly totals) — cross-sheet ref
    # ══════════════════════════════════════════════════════════════════════════
    ws_mr = wb["Monthly Report"]

    section_hdr(ws_mr,
        "MONTHLY EARNED VALUE ANALYSIS  ·  Data sourced from DASHBOARD SUMIFS (cross-sheet reference)",
        "A22:O22", height=22)

    # Clustered column — EV vs PV
    c4 = make_bar("Earned Value vs Planned Value by Month (NOK)",
                  chart_type="col", y_fmt='#,##0', w=20, h=12)
    c4.y_axis.title = "NOK"

    cats4   = Reference(ws_dash, min_col=1, min_row=29, max_row=31)  # months
    ev_data = Reference(ws_dash, min_col=5, min_row=28, max_row=31)  # EV + header
    pv_data = Reference(ws_dash, min_col=6, min_row=28, max_row=31)  # PV + header
    c4.add_data(ev_data, titles_from_data=True)
    c4.add_data(pv_data, titles_from_data=True)
    bar_line(c4.series[0], BLUE)    # EV — accent blue
    bar_line(c4.series[1], SLATE)   # PV — muted slate
    c4.set_categories(cats4)
    ws_mr.add_chart(c4, "A23")

    # Line chart — SV% (Schedule Variance) trend by month
    c5 = LineChart()
    c5.style  = 10
    c5.title  = "Schedule Variance % by Month  ·  Target ≥ 0%"
    c5.y_axis.numFmt = "+0.0%;-0.0%;0.0%"
    c5.y_axis.title  = "Schedule Variance %"
    c5.width  = 14
    c5.height = 12
    c5.legend.position = "b"

    cats5   = Reference(ws_dash, min_col=1, min_row=29, max_row=31)
    # Build a tiny helper column in Monthly Report with per-month SV%
    # Use Dashboard's Hours variance as a proxy (col 4 = Actual-Budget)
    sv_data = Reference(ws_dash, min_col=8, min_row=28, max_row=31)  # CPI header+data
    c5.add_data(sv_data, titles_from_data=True)
    line_style(c5.series[0], BLUE, width_emu=28000)
    c5.set_categories(cats5)
    ws_mr.add_chart(c5, "K23")

    print("Charts added: Discipline Progress, SPI, Activity Progress, EV/PV Trend, CPI Line")


if __name__ == "__main__":
    main()
